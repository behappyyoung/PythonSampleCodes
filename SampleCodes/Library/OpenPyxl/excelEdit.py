# from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook(filename="KAPA.xltx")
s_names = workbook.sheetnames
a_sheet = workbook.active

print(s_names)
print(a_sheet.title)
s_sheet = workbook.worksheets[3]
print(s_sheet.title)
for row in s_sheet.iter_rows(min_row=4, max_row=20, min_col=4, max_col=8, values_only=True):
    print(row)
print(s_sheet["F80"].value)
s_sheet["F80"] = 8.88

workbook.save(filename='KAPA.xltx')
